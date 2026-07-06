"""
ARMS Bulk Student Data Fetcher
==============================
Logs in as faculty, fetches data for ALL students, and exports to Excel.
Usage: python arms_all_students.py
Output: arms_students_data.xlsx
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import requests
from bs4 import BeautifulSoup
import json, time, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

BASE = "https://arms.sse.saveetha.com"

# ── Credentials ──────────────────────────────────────────────────────────────
FACULTY_USER = "SSETSCS262"
FACULTY_PASS = "kumbakonam123$"

OUTPUT_FILE = "arms_students_data.xlsx"

# ── Session ──────────────────────────────────────────────────────────────────
session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
})

# ── Helpers ──────────────────────────────────────────────────────────────────
def login_faculty():
    print("\n[*] Logging in as faculty...")
    r = session.get(BASE + "/")
    soup = BeautifulSoup(r.text, "html.parser")
    payload = {
        "__VIEWSTATE":          soup.find("input", {"name": "__VIEWSTATE"})["value"],
        "__VIEWSTATEGENERATOR": soup.find("input", {"name": "__VIEWSTATEGENERATOR"})["value"],
        "__EVENTVALIDATION":    soup.find("input", {"name": "__EVENTVALIDATION"})["value"],
        "txtusername":          FACULTY_USER,
        "txtpassword":          FACULTY_PASS,
        "btnlogin":             "Login",
    }
    r2 = session.post(BASE + "/", data=payload, allow_redirects=True)
    if "FacultyPortal" in r2.url:
        print(f"    [OK] Faculty login successful -> {r2.url}")
        return True
    print(f"    [FAIL] Login failed -> {r2.url}")
    return False

def api(handler, page, mode, extra=None):
    url = f"{BASE}/Handler/{handler}.ashx"
    params = {"Page": page, "Mode": mode}
    if extra:
        params.update(extra)
    try:
        r = session.get(url, params=params, timeout=15)
        if r.text.strip():
            return r.json()
        return {}
    except Exception:
        return {}

def get_all_students():
    """Fetch the full student list using GETALLRECORDREGNOLIBS."""
    print("\n[*] Fetching student list...")
    data = api("Student", "StudentView", "GETALLRECORDREGNOLIBS", {"Id": ""})
    if "Table" in data and data["Table"]:
        students = data["Table"]
        print(f"    [OK] Found {len(students)} students")
        return students
    # Try alternate
    print("    [!] Trying alternate student list endpoint...")
    data2 = api("Administration", "PrincDashInstitute", "StudentDetailsById", {"Id": ""})
    if "Table" in data2:
        print(f"    [OK] Found {len(data2['Table'])} students (alt)")
        return data2["Table"]
    print("    [WARN] No students found via auto-discovery.")
    return []

def get_student_results(student_id):
    data = api("Administration", "PRINCGETENROLLCOURSE", "GETRESULT", {"Id": student_id})
    if isinstance(data, list):
        return data
    return data.get("Table", []) if isinstance(data, dict) else []

def get_student_attendance(student_id):
    data = api("Administration", "CourseDateByProgramStuDean", "ATTENDANCESTUPERSENT", {"SId": student_id})
    if isinstance(data, list):
        return data
    return data.get("Table", []) if isinstance(data, dict) else []

def get_student_enrollment(student_id):
    data = api("Parents", "StudentDetails", "CourseCompleteStatus", {"StudentId": student_id})
    if isinstance(data, list):
        return data[0] if data else {}
    rows = data.get("Table", []) if isinstance(data, dict) else []
    return rows[0] if rows else {}

def get_student_payment(student_id):
    data = api("Parents", "StudentDetails", "Paymentlist", {"StudentId": student_id})
    if isinstance(data, list):
        return data
    return data.get("Table", []) if isinstance(data, dict) else []

# ── Excel Styling ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
PASS_FILL     = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL     = PatternFill("solid", fgColor="FFC7CE")
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
WARN_FILL     = PatternFill("solid", fgColor="FFEB9C")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
PASS_FONT     = Font(name="Calibri", color="375623", bold=True, size=10)
FAIL_FONT     = Font(name="Calibri", color="9C0006", bold=True, size=10)

thin_border = Border(
    left=Side(style='thin', color='BDD7EE'),
    right=Side(style='thin', color='BDD7EE'),
    top=Side(style='thin', color='BDD7EE'),
    bottom=Side(style='thin', color='BDD7EE')
)

def style_header(cell, text=None):
    if text:
        cell.value = text
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

def style_subheader(cell, text=None):
    if text:
        cell.value = text
    cell.fill   = SUBHEAD_FILL
    cell.font   = SUBHEAD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

def style_body(cell, alt=False):
    cell.fill   = ALT_FILL if alt else WHITE_FILL
    cell.font   = BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border

def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))

# ── Sheet Builders ────────────────────────────────────────────────────────────
def build_summary_sheet(wb, all_data):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20

    # Title row
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = f"ARMS — All Students Academic Report   |   Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}"
    title.fill  = PatternFill("solid", fgColor="0D47A1")
    title.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Reg No.", "Student Name", "Department", "Program",
               "Batch", "Total Courses", "Passed", "Failed",
               "Pass %", "CGPA (Grade)", "Arrear Subjects"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(2, ci), h)

    ws.freeze_panes = "A3"

    for ri, sd in enumerate(all_data, 1):
        info     = sd.get("info", {})
        results  = sd.get("results", [])
        passed   = [r for r in results if r.get("FinalResult") == "PASS"]
        failed   = [r for r in results if r.get("FinalResult") == "FAIL"]
        total    = len(results)
        pass_pct = round(len(passed) / total * 100, 1) if total else 0
        arrears  = ", ".join([r.get("CourseName", r.get("CourseCode", "?")) for r in failed]) or "—"

        grades   = [r.get("FinalGrade", "") for r in results if r.get("FinalGrade")]
        grade_summary = ", ".join(sorted(set(grades))) if grades else "—"

        row_data = [
            ri,
            info.get("RegNumber", info.get("StudentId", sd.get("id", ""))),
            info.get("StudentName", info.get("Name", "—")),
            info.get("Department", info.get("DeptName", "—")),
            info.get("Program",    info.get("ProgramName", "—")),
            info.get("Batch",      info.get("BatchYear", "—")),
            total, len(passed), len(failed),
            f"{pass_pct}%",
            grade_summary,
            arrears,
        ]
        alt = (ri % 2 == 0)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(ri + 2, ci)
            cell.value = val
            style_body(cell, alt)
            # Colour pass%
            if ci == 10:
                if pass_pct == 100:
                    cell.fill = PASS_FILL; cell.font = PASS_FONT
                elif pass_pct < 80:
                    cell.fill = FAIL_FILL; cell.font = FAIL_FONT
                else:
                    cell.fill = WARN_FILL
            # Colour failed count
            if ci == 9 and len(failed) > 0:
                cell.fill = FAIL_FILL; cell.font = FAIL_FONT

    auto_width(ws)
    ws.column_dimensions["L"].width = 45  # Arrear subjects wider
    return ws

def build_results_sheet(wb, all_data):
    ws = wb.create_sheet("All Results")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "ALL STUDENTS — DETAILED EXAM RESULTS"
    t.fill  = PatternFill("solid", fgColor="0D47A1")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Reg No.", "Student Name", "S.No", "Course Code", "Course Name",
               "Grade", "Result", "Max Marks", "Month/Year", "Remarks"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(2, ci), h)

    ws.freeze_panes = "A3"
    row = 3
    for sd in all_data:
        info    = sd.get("info", {})
        results = sd.get("results", [])
        reg     = info.get("RegNumber", info.get("StudentId", sd.get("id", "")))
        name    = info.get("StudentName", info.get("Name", "—"))
        for sno, r in enumerate(results, 1):
            result_val = r.get("FinalResult", "")
            is_pass    = result_val == "PASS"
            is_fail    = result_val == "FAIL"
            month      = r.get("MonthYearValue", "")[:10] if r.get("MonthYearValue") else ""
            row_data   = [
                reg, name, sno,
                r.get("CourseCode", ""),
                r.get("CourseName", ""),
                r.get("FinalGrade", ""),
                result_val,
                r.get("MaxMark", ""),
                month,
                "Arrear" if is_fail else "",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row, ci)
                cell.value = val
                if is_fail:
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT
                elif is_pass:
                    cell.fill = PASS_FILL if ci in (6, 7) else (ALT_FILL if row % 2 == 0 else WHITE_FILL)
                    if ci in (6, 7):
                        cell.font = PASS_FONT
                    else:
                        cell.font = BODY_FONT
                else:
                    cell.fill = ALT_FILL if row % 2 == 0 else WHITE_FILL
                    cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border
            row += 1
        # Blank separator row between students
        row += 1

    auto_width(ws)
    return ws

def build_attendance_sheet(wb, all_data):
    ws = wb.create_sheet("Attendance")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "ALL STUDENTS — ATTENDANCE REPORT"
    t.fill  = PatternFill("solid", fgColor="0D47A1")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Reg No.", "Student Name", "Course Code", "Subject Name",
               "Slot", "Faculty", "Absent", "Total Classes"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(2, ci), h)

    ws.freeze_panes = "A3"
    row = 3
    for sd in all_data:
        info = sd.get("info", {})
        att  = sd.get("attendance", [])
        reg  = info.get("RegNumber", info.get("StudentId", sd.get("id", "")))
        name = info.get("StudentName", info.get("Name", "—"))
        if att:
            for a in att:
                absent = a.get("AbsentCount", 0)
                total  = a.get("Totalclasscount", 0)
                row_data = [
                    reg, name,
                    a.get("SubjectCode", ""),
                    a.get("SubjectName", ""),
                    a.get("SlotId", ""),
                    a.get("FacultyName", ""),
                    absent,
                    total,
                ]
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row, ci)
                    cell.value = val
                    style_body(cell, row % 2 == 0)
                row += 1
        else:
            # No attendance data row
            for ci, val in enumerate([reg, name, "—", "(No attendance data)", "", "", "", ""], 1):
                cell = ws.cell(row, ci)
                cell.value = val
                style_body(cell, row % 2 == 0)
            row += 1
        row += 1  # separator

    auto_width(ws)
    return ws

def build_arrears_sheet(wb, all_data):
    ws = wb.create_sheet("Arrears Only")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "STUDENTS WITH ARREARS (FAILED SUBJECTS)"
    t.fill  = PatternFill("solid", fgColor="B71C1C")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Reg No.", "Student Name", "Course Code", "Failed Subject", "Grade", "Exam Month", "Arrear Count"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(2, ci), h)

    ws.freeze_panes = "A3"
    row = 3
    for sd in all_data:
        info    = sd.get("info", {})
        results = sd.get("results", [])
        failed  = [r for r in results if r.get("FinalResult") == "FAIL"]
        if not failed:
            continue
        reg  = info.get("RegNumber", info.get("StudentId", sd.get("id", "")))
        name = info.get("StudentName", info.get("Name", "—"))
        for r in failed:
            month = r.get("MonthYearValue", "")[:10] if r.get("MonthYearValue") else ""
            row_data = [
                reg, name,
                r.get("CourseCode", ""),
                r.get("CourseName", ""),
                r.get("FinalGrade", ""),
                month,
                len(failed),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row, ci)
                cell.value = val
                cell.fill  = FAIL_FILL
                cell.font  = FAIL_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border
            row += 1

    auto_width(ws)
    return ws

# ── Main ──────────────────────────────────────────────────────────────────────
def fetch_one(stu):
    """Fetch all data for a single student (used by thread pool)."""
    sid  = (stu.get("RegNumber") or stu.get("StudentId") or
            stu.get("Id") or stu.get("id") or "")
    name = (stu.get("StudentName") or stu.get("Name") or
            stu.get("StudentFullName") or sid)
    results    = get_student_results(sid)
    attendance = get_student_attendance(sid)
    enrollment = get_student_enrollment(sid)
    passed = len([r for r in results if r.get("FinalResult") == "PASS"])
    failed = len([r for r in results if r.get("FinalResult") == "FAIL"])
    return {
        "sid": sid, "name": name,
        "id":         sid,
        "info":       {**stu, **enrollment},
        "results":    results,
        "attendance": attendance,
        "_passed":    passed,
        "_failed":    failed,
    }

def main():
    import concurrent.futures, threading

    if not login_faculty():
        sys.exit(1)

    students = get_all_students()
    if not students:
        print("\n[!] Fallback: using only the sample student.")
        students = [{"RegNumber": "192411184", "StudentName": "Student"}]

    total = len(students)

    # ── Inspect available filters ─────────────────────────────────────────────
    sample = students[:5]
    keys   = set()
    for s in sample:
        keys.update(s.keys())
    print(f"\n  Available student fields: {sorted(keys)}")

    # Detect department / batch / program fields
    dept_key  = next((k for k in ["DeptName","Department","Dept","dept"] if k in keys), None)
    batch_key = next((k for k in ["Batch","BatchYear","batch","AcademicYear"] if k in keys), None)
    prog_key  = next((k for k in ["ProgramName","Program","programme"] if k in keys), None)
    reg_key   = next((k for k in ["RegNumber","StudentId","Id","id"] if k in keys), None)

    if dept_key:
        depts = sorted(set(s.get(dept_key, "") for s in students if s.get(dept_key)))
        print(f"\n  Departments ({len(depts)}): {', '.join(depts[:20])}")
    if batch_key:
        batches = sorted(set(str(s.get(batch_key, "")) for s in students if s.get(batch_key)))
        print(f"  Batches     ({len(batches)}): {', '.join(batches[:20])}")

    print(f"\n  Total students in portal: {total}")
    print("─"*60)
    print("  FILTER OPTIONS (press Enter to skip a filter):")

    # Filter by reg prefix (e.g. '192' for 2019 batch)
    reg_prefix = input("  Filter by Reg No. prefix (e.g. 192, 202, blank=all): ").strip()
    # Filter by department
    dept_filter = ""
    if dept_key:
        dept_filter = input(f"  Filter by dept name contains (e.g. CSE, blank=all): ").strip().upper()
    # Limit
    limit_str = input(f"  Max students to process (blank=ALL {total}): ").strip()
    limit = int(limit_str) if limit_str.isdigit() else total

    # Apply filters
    filtered = students
    if reg_prefix:
        filtered = [s for s in filtered if str(s.get(reg_key, "")).startswith(reg_prefix)]
    if dept_filter and dept_key:
        filtered = [s for s in filtered if dept_filter in str(s.get(dept_key, "")).upper()]
    filtered = filtered[:limit]

    print(f"\n  [*] Will process {len(filtered)} students (threads=8)...\n")
    if not filtered:
        print("  No students matched filters. Exiting.")
        sys.exit(0)

    # ── Concurrent fetch ──────────────────────────────────────────────────────
    all_data = []
    lock = threading.Lock()
    done = [0]

    def process(stu):
        result = fetch_one(stu)
        with lock:
            done[0] += 1
            n = done[0]
            print(f"  [{n:>5}/{len(filtered)}] {result['sid']:<15} {result['name']:<35}"
                  f" {result['_passed']}P/{result['_failed']}F")
        return result

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        futures = [ex.submit(process, s) for s in filtered]
        for f in concurrent.futures.as_completed(futures):
            try:
                all_data.append(f.result())
            except Exception as e:
                print(f"  [ERR] {e}")

    # Sort by reg number for clean output
    all_data.sort(key=lambda x: str(x.get("id", "")))

    # ── Build Excel ──────────────────────────────────────────────────────────
    print(f"\n[*] Building Excel report for {len(all_data)} students...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, all_data)
    build_results_sheet(wb, all_data)
    build_attendance_sheet(wb, all_data)
    build_arrears_sheet(wb, all_data)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = f"arms_students_{ts}.xlsx"
    wb.save(out)

    failed_students = sum(1 for sd in all_data
                          if any(r.get("FinalResult") == "FAIL"
                                 for r in sd.get("results", [])))
    print(f"\n{'='*60}")
    print(f"  [OK] Report saved -> {os.path.abspath(out)}")
    print(f"  Students processed : {len(all_data)}")
    print(f"  Students w/ arrears: {failed_students}")
    print(f"  Sheets: Summary | All Results | Attendance | Arrears Only")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
